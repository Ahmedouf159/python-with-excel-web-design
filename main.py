from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from config import Config
from models import db, Student
from excel_processor import process_excel_files, generate_unique_line_code
from datetime import datetime
import qrcode
import os
from fpdf import FPDF
from openpyxl import Workbook
import pandas as pd
import random

app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = 'secret_key_123'
ADMIN_SECRET_KEY = "admin123"
QUIZ_SECRET = "quiz123"
db.init_app(app)

tips = [
    "Review your lessons regularly.",
    "Take breaks while studying.",
    "Sleep well before exams.",
    "Use diagrams to understand better.",
    "Ask your teacher when confused."
]
questions = [
    {"q": "What is 2 + 2?", "a": "4"},
    {"q": "What is the capital of Egypt?", "a": "cairo"},
    {"q": "What is 5 x 3?", "a": "15"}
]

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        input_code = request.form.get('secret_key', '').strip()
        with open("log.txt", "a", encoding="utf-8") as log:
            log.write(f"\U0001f510 Login attempt with code: {input_code} - {datetime.now()}\n")

        if input_code == "admin123":
            session['logged_in'] = True
            with open("log.txt", "a", encoding="utf-8") as log:
                log.write(f"\u2705 Admin login - {datetime.now()}\n")
            return redirect(url_for('dashboard'))

        elif input_code == "quiz123":
            session['challenge_mode'] = True
            session['question_index'] = 0
            with open("log.txt", "a", encoding="utf-8") as log:
                log.write(f"\u2705 Challenge login - {datetime.now()}\n")
            return redirect(url_for('challenge_mode'))

        else:
            flash("\u274c Incorrect secret code.")
            with open("log.txt", "a", encoding="utf-8") as log:
                log.write(f"\u274c Failed login - {datetime.now()}\n")

    return render_template('login.html')


# CHALLENGE MODE ROUTE
@app.route('/challenge', methods=['GET', 'POST'])
def challenge_mode():
    if not session.get('challenge_mode'):
        return redirect(url_for('login'))

    index = session.get('question_index', 0)

    if index >= len(questions):
        flash("\U0001f389 You completed the challenge!")
        session.pop('challenge_mode', None)
        session.pop('question_index', None)
        return redirect(url_for('login'))

    question = questions[index]

    if request.method == 'POST':
        user_answer = request.form.get('answer', '').strip().lower()
        correct_answer = question['a'].strip().lower()

        if user_answer == correct_answer:
            session['question_index'] = index + 1
            flash("\u2705 Correct! Next question...")
        else:
            flash("\u274c Incorrect, try again.")

        return redirect(url_for('challenge_mode'))

    return render_template(
        'challenge.html',
        question=question['q'],
        index=index + 1,
        total=len(questions)
    )


@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    flash("🔒 Logged out.")
    return redirect(url_for('login'))

@app.route('/dashboard')
def dashboard():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    sort_by = request.args.get('sort', 'name')
    order = request.args.get('order', 'asc')
    query = request.args.get('query', '')

    q = Student.query
    if query:
        q = q.filter(Student.name.ilike(f'%{query}%'))

    students = q.order_by(
        getattr(Student, sort_by).asc() if order == 'asc' else getattr(Student, sort_by).desc()
    ).all()

    for student in students:
        student.total = student.math + student.english + student.science

    top_student = max(students, key=lambda s: s.total, default=None)
    low_student = min(students, key=lambda s: s.total, default=None)
    last_updated = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    daily_tip = random.choice(tips)

    return render_template('dashboard.html', students=students, query=query, order=order,
                           sort_by=sort_by, last_updated=last_updated,
                           top_student=top_student, low_student=low_student, tip=daily_tip)

@app.route('/ai_query', methods=['POST'])
def ai_query():
    keyword = request.form.get('query')
    if 'low' in keyword.lower() and 'english' in keyword.lower():
        results = Student.query.filter(Student.english < 50).all()
        return jsonify([{"name": s.name, "english": s.english} for s in results])
    return jsonify([])

@app.route('/add_student', methods=['POST'])
def add_student():
    name = request.form['name']
    math = request.form['math']
    english = request.form['english']
    science = request.form['science']
    address = request.form['address']

    with open("log.txt", "a", encoding="utf-8") as log:
        log.write(f"✅ Added: {name} - {datetime.now()}\n")

    exists = Student.query.filter_by(name=name).first()
    if exists:
        flash("⚠️ Student already exists.")
    else:
        student = Student(
            name=name, math=math, english=english, science=science,
            address=address, line_code=generate_unique_line_code(),
            created_at=datetime.now()
        )
        db.session.add(student)
        db.session.commit()
        export_to_merged_excel()
        flash("✅ Student added.")
    return redirect(url_for('dashboard'))

@app.route('/edit/<int:student_id>', methods=['GET', 'POST'])
def edit_student(student_id):
    student = Student.query.get_or_404(student_id)
    if request.method == 'POST':
        student.name = request.form['name']
        student.math = int(request.form['math'])
        student.english = int(request.form['english'])
        student.science = int(request.form['science'])
        student.address = request.form['address']
        db.session.commit()
        with open("log.txt", "a", encoding="utf-8") as log:
            log.write(f"✏️ Edited: {student.name} - {datetime.now()}\n")
        flash(f"✏️ Updated {student.name}")
        return redirect(url_for('dashboard'))
    return render_template('edit_student.html', student=student)



@app.route('/delete_student/<int:student_id>', methods=['POST'])
def delete_student(student_id):
    student = Student.query.get(student_id)
    if student:
        db.session.delete(student)
        db.session.commit()
        with open("log.txt", "a", encoding="utf-8") as log:
            log.write(f"🗑️ Deleted: {student.name} - {datetime.now()}\n")
        export_to_merged_excel()
        flash(f"🗑️ {student.name} deleted.")
    return redirect(url_for('dashboard'))

@app.route('/import_excel', methods=['POST'])
def import_excel():
    process_excel_files(
        'i:/Ahmed/websites/python with excel/student.xlsx',
        'i:/Ahmed/websites/python with excel/add.xlsx'
    )
    export_to_merged_excel()
    flash("📁 Excel imported.")
    return redirect(url_for('dashboard'))

@app.route('/save_to_excel', methods=['POST'])
def save_to_excel():
    export_to_merged_excel()
    with open("log.txt", "a", encoding="utf-8") as log:
        log.write(f"💾 Saved to Excel - {datetime.now()}\n")
    flash("💾 Excel saved.")
    return redirect(url_for('dashboard'))

@app.route('/qr/<line_code>')
def generate_qr(line_code):
    img = qrcode.make(line_code)
    path = f'static/qrcodes/{line_code}.png'
    os.makedirs('static/qrcodes', exist_ok=True)
    img.save(path)
    return send_file(path, mimetype='image/png')

@app.route('/export/excel')
def export_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(['ID', 'Name', 'Math', 'English', 'Science', 'Address', 'Line Code'])
    for s in Student.query.all():
        ws.append([s.id, s.name, s.math, s.english, s.science, s.address, s.line_code])
    path = "static/exports/students.xlsx"
    wb.save(path)
    return send_file(path, as_attachment=True)

@app.route('/export/<subject>')
def export_by_subject(subject):
    if subject not in ['math', 'english', 'science']:
        flash("❌ Invalid subject.")
        return redirect(url_for('dashboard'))

    wb = Workbook()
    ws = wb.active
    ws.title = subject.capitalize()
    ws.append(['ID', 'Name', subject.capitalize()])
    for s in Student.query.all():
        score = getattr(s, subject)
        ws.append([s.id, s.name, score])
    path = f"static/exports/{subject}_students.xlsx"
    wb.save(path)
    return send_file(path, as_attachment=True)

@app.route('/logs')
def view_logs():
    if not os.path.exists("log.txt"):
        return "<h4>No logs available.</h4><a href='/dashboard'>Back</a>"
    with open("log.txt", "r", encoding="utf-8") as f:
        logs = f.read().splitlines()
    return "<h4>🧾 Operation Logs:</h4><ul>" + "".join(f"<li>{log}</li>" for log in logs) + "</ul><a href='/dashboard'>⬅ Back</a>"

@app.route('/send_email', methods=['GET', 'POST'])
def send_email_form():
    if request.method == 'POST':
        email = request.form['email']
        flash(f"📧 Simulated sending Excel to {email}")
        with open("log.txt", "a", encoding="utf-8") as log:
            log.write(f"📧 Sent email to {email} - {datetime.now()}\n")
        return redirect(url_for('dashboard'))
    return """
    <h3>📧 Send Excel via Email</h3>
    <form method='post'>
        <input type='email' name='email' placeholder='Enter email' required>
        <button type='submit'>Send</button>
    </form>
    <a href='/dashboard'>⬅ Back</a>
    """

@app.route('/export/pdf')
def export_pdf():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    for s in Student.query.all():
        text = f"{s.name} | Math: {s.math}, Eng: {s.english}, Sci: {s.science}"
        pdf.cell(200, 10, txt=text, ln=True)
    path = "static/exports/students.pdf"
    pdf.output(path)
    return send_file(path, as_attachment=True)

@app.route('/clear_data', methods=['POST'])
def clear_data():
    db.session.query(Student).delete()
    db.session.commit()
    export_to_merged_excel()
    flash("🗑️ All data deleted.")
    with open("log.txt", "a", encoding="utf-8") as log:
        log.write(f"🗑️ All data cleared - {datetime.now()}\n")
    return redirect(url_for('dashboard'))

def export_to_merged_excel():
    data = [{
        "StudentName": s.name,
        "Math": s.math,
        "English": s.english,
        "Science": s.science,
        "Address": s.address,
        "line code": s.line_code
    } for s in Student.query.all()]
    df = pd.DataFrame(data)
    df.to_excel(r"I:/Ahmed/websites/python with excel/merged_output.xlsx", index=False)

def create_database():
    with app.app_context():
        db.create_all()
        print("✅ Database created.")

def generate_data():
    with app.app_context():
        process_excel_files(
            'i:/Ahmed/websites/python with excel/student.xlsx',
            'i:/Ahmed/websites/python with excel/add.xlsx'
        )
        export_to_merged_excel()

def run_website():
    app.run(debug=True)

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python main.py [initdb|generate|runserver]")
    elif sys.argv[1] == "initdb":
        create_database()
    elif sys.argv[1] == "generate":
        generate_data()
    elif sys.argv[1] == "runserver":
        run_website()

